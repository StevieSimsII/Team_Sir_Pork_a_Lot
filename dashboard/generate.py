"""
dashboard/generate.py

Fetches raffle data (all years) from SharePoint via Microsoft Graph API
and renders a self-contained neon-themed HTML dashboard to dashboard/dist/index.html.

Includes:
  - Current-year stats, goal bar, daily sales chart, tier donut, top 5
  - Year-over-year comparison card (revenue/tickets/buyers through same date)
  - Cumulative sales line chart: full prior year vs current year to date
  - Returning buyers table  (bought last year AND this year)
  - Prospects table         (bought last year, NOT yet this year — solicit these)
  - New buyers table        (first-timers this year)
"""

import csv
import html as html_lib
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config from environment ───────────────────────────────────────────────────

TENANT_ID     = os.environ["AZURE_TENANT_ID"]
CLIENT_ID     = os.environ["AZURE_CLIENT_ID"]
CLIENT_SECRET = os.environ["AZURE_CLIENT_SECRET"]
SP_HOSTNAME   = os.environ.get("SHAREPOINT_HOSTNAME", "kingsofcode.sharepoint.com")
SP_SITE_PATH  = os.environ.get("SHAREPOINT_SITE_PATH", "/sites/StevieCopilot")
LIST_ID       = os.environ["SHAREPOINT_LIST_ID"]
GOAL          = int(os.environ.get("GOAL_AMOUNT", "30000"))
CRNA_GOAL     = int(os.environ.get("CRNA_BUNDLE_GOAL_AMOUNT", "0"))

_TZ          = ZoneInfo("America/Chicago")
TODAY        = datetime.now(_TZ).date()
CURRENT_YEAR = TODAY.year
PREV_YEAR    = CURRENT_YEAR - 1
SAME_DAY_PREV = date(PREV_YEAR, TODAY.month, TODAY.day)

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH    = os.path.normpath(
    os.path.join(_SCRIPT_DIR, "..", "Hogs for the Cause 2025 Raffle.csv")
)

# ── Auth ──────────────────────────────────────────────────────────────────────

def get_token() -> str:
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    resp = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope":         "https://graph.microsoft.com/.default",
    }, timeout=30)
    resp.raise_for_status()
    return resp.json()["access_token"]

def graph(token: str, url: str) -> dict:
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
    resp.raise_for_status()
    return resp.json()

# ── Data fetch ────────────────────────────────────────────────────────────────

def get_site_id(token: str) -> str:
    data = graph(token, f"https://graph.microsoft.com/v1.0/sites/{SP_HOSTNAME}:{SP_SITE_PATH}")
    return data["id"]

def get_list_items(token: str, site_id: str) -> list[dict]:
    """Fetch ALL items (all years) with pagination."""
    url = (
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{LIST_ID}/items"
        f"?$expand=fields&$top=999"
    )
    all_items = []
    while url:
        data = graph(token, url)
        all_items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    print(f"  Total items in list: {len(all_items)}")
    return all_items


# ── Normalisation helpers ─────────────────────────────────────────────────────

def norm_name(n: str) -> str:
    return " ".join(str(n or "").lower().split())

def norm_phone(p: str) -> str:
    digits = re.sub(r"\D", "", str(p or ""))
    return digits[-10:] if len(digits) >= 10 else digits

def norm_email(e: str) -> str:
    return str(e or "").lower().strip()

def esc(s: str) -> str:
    return html_lib.escape(str(s or ""))


# ── CSV loader (2025 supplemental data) ──────────────────────────────────────

def load_csv_prev() -> list[dict]:
    rows: list[dict] = []
    try:
        with open(CSV_PATH, newline="", encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                name  = (row.get("Name")          or "").strip()
                email = (row.get("Email Address") or "").strip()
                phone = (row.get("Phone Number")  or "").strip()
                if name:
                    rows.append({"name": name, "email": email, "phone": phone})
    except FileNotFoundError:
        print(f"  WARNING: CSV not found at {CSV_PATH}", file=sys.stderr)
    print(f"  CSV {PREV_YEAR} rows: {len(rows)}")
    return rows


# ── Buyer record ──────────────────────────────────────────────────────────────

class Buyer:
    __slots__ = ("name", "email", "phone",
                 "tickets_prev", "amount_prev",
                 "tickets_cur",  "amount_cur")
    def __init__(self, name="", email="", phone=""):
        self.name         = name
        self.email        = email
        self.phone        = phone
        self.tickets_prev = 0
        self.amount_prev  = 0.0
        self.tickets_cur  = 0
        self.amount_cur   = 0.0


RAFFLE_CONFIGS = {
    "hogs_for_the_cause": {
        "slug": "hogs",
        "display_name": "Hogs for the Cause",
        "short_name": "Hogs",
        "goal": GOAL,
        "include_history": True,
    },
    "crna_essential_bundle": {
        "slug": "crna",
      "display_name": "APEX CRNA Essential Bundle",
      "short_name": "APEX CRNA Bundle",
        "goal": CRNA_GOAL,
        "include_history": False,
    },
}

# ── Processing ────────────────────────────────────────────────────────────────

def process_raffle(all_items: list[dict], raffle_key: str, csv_rows: list[dict] | None = None) -> dict:
  config = RAFFLE_CONFIGS[raffle_key]
  include_history = config["include_history"]

  cur_fields: list[dict] = []
  prev_fields: list[dict] = []
  for item in all_items:
    fields = item.get("fields", {})
    raffle = (fields.get("RaffleName") or "").strip().lower()
    if raffle != raffle_key:
      continue
    year = (fields.get("SubmissionDate") or "")[:4]
    if year == str(CURRENT_YEAR):
      cur_fields.append(fields)
    elif include_history and year == str(PREV_YEAR):
      prev_fields.append(fields)

  print(f"  {config['display_name']}: {len(cur_fields)} current-year items")
  if include_history:
    print(f"    {PREV_YEAR}: {len(prev_fields)} prior-year items")

  def _key(email: str, phone: str, name: str) -> str:
    e = norm_email(email)
    p = norm_phone(phone)
    n = norm_name(name)
    return e if e else (p if p else n)

  total_raised = 0.0
  total_tickets = 0
  buyers_cur: set[str] = set()
  daily_totals: dict[str, float] = defaultdict(float)
  tier_counts: dict[int, int] = defaultdict(int)
  top_buyers: dict[str, float] = defaultdict(float)
  map_cur: dict[str, Buyer] = {}
  daily_cur: dict[str, float] = defaultdict(float)

  for fields in cur_fields:
    tickets = int(fields.get("NumberofChances", 0) or 0)
    if tickets == 0:
      continue
    amount = float(fields.get("TotalPaid", 0) or 0)
    person = str(fields.get("Person", "Unknown") or "Unknown").strip()
    email = str(fields.get("Email", "") or "").strip()
    phone = str(fields.get("Phone", "") or "").strip()
    day = (fields.get("SubmissionDate") or "")[:10]

    total_raised += amount
    total_tickets += tickets
    buyers_cur.add(person.lower())
    top_buyers[person] += amount
    tier_counts[tickets] += 1
    if day:
      daily_totals[day] += amount
      daily_cur[day] += amount

    buyer_key = _key(email, phone, person)
    if buyer_key not in map_cur:
      map_cur[buyer_key] = Buyer(name=person, email=email, phone=phone)
    map_cur[buyer_key].tickets_cur += tickets
    map_cur[buyer_key].amount_cur += amount

  sorted_days = sorted(daily_totals.keys())
  top5 = sorted(top_buyers.items(), key=lambda item: item[1], reverse=True)[:5]
  tier_label_map = {1: "1 Ticket ($25)", 3: "3 Tickets ($60)", 6: "6 Tickets ($100)", 12: "12 Tickets ($200)"}
  tiers = [(tier_label_map.get(count, f"{count} Tickets"), total) for count, total in sorted(tier_counts.items())]

  map_prev: dict[str, Buyer] = {}
  daily_prev: dict[str, float] = defaultdict(float)
  todate_prev_amt = 0.0
  todate_prev_tix = 0
  todate_prev_keys: set[str] = set()

  for fields in prev_fields:
    amount = float(fields.get("TotalPaid", 0) or 0)
    tickets = int(fields.get("NumberofChances", 0) or 0)
    person = str(fields.get("Person", "Unknown") or "Unknown").strip()
    email = str(fields.get("Email", "") or "").strip()
    phone = str(fields.get("Phone", "") or "").strip()
    day = (fields.get("SubmissionDate") or "")[:10]

    if day:
      daily_prev[day] += amount
      if day <= str(SAME_DAY_PREV):
        todate_prev_amt += amount
        todate_prev_tix += tickets

    buyer_key = _key(email, phone, person)
    if day and day <= str(SAME_DAY_PREV):
      todate_prev_keys.add(buyer_key)
    if buyer_key not in map_prev:
      map_prev[buyer_key] = Buyer(name=person, email=email, phone=phone)
    map_prev[buyer_key].tickets_prev += tickets
    map_prev[buyer_key].amount_prev += amount

  if include_history and csv_rows:
    for row in csv_rows:
      email_key = norm_email(row["email"])
      phone_key = norm_phone(row["phone"])
      name_key = norm_name(row["name"])
      matched_key = None

      if email_key and email_key in map_prev:
        matched_key = email_key
      elif phone_key and phone_key in map_prev:
        matched_key = phone_key
      else:
        for existing_key, buyer in map_prev.items():
          if norm_name(buyer.name) == name_key:
            matched_key = existing_key
            break

      if matched_key:
        buyer = map_prev[matched_key]
        if not buyer.email and row["email"]:
          buyer.email = row["email"]
        if not buyer.phone and row["phone"]:
          buyer.phone = row["phone"]
      else:
        fallback_key = email_key or phone_key or name_key
        if fallback_key and fallback_key not in map_prev:
          map_prev[fallback_key] = Buyer(name=row["name"], email=row["email"], phone=row["phone"])

  emails_cur = {norm_email(buyer.email) for buyer in map_cur.values() if buyer.email}
  phones_cur = {norm_phone(buyer.phone) for buyer in map_cur.values() if buyer.phone}
  names_cur = {norm_name(buyer.name) for buyer in map_cur.values()}
  emails_prev = {norm_email(buyer.email) for buyer in map_prev.values() if buyer.email}
  phones_prev = {norm_phone(buyer.phone) for buyer in map_prev.values() if buyer.phone}
  names_prev = {norm_name(buyer.name) for buyer in map_prev.values()}

  returning: list[Buyer] = []
  prospects: list[Buyer] = []
  for buyer in map_prev.values():
    email_key = norm_email(buyer.email)
    phone_key = norm_phone(buyer.phone)
    name_key = norm_name(buyer.name)
    is_returning = (
      (email_key and email_key in emails_cur)
      or (phone_key and phone_key in phones_cur)
      or (name_key and name_key in names_cur)
    )
    if is_returning:
      for current_buyer in map_cur.values():
        email_match = email_key and norm_email(current_buyer.email) == email_key
        phone_match = phone_key and norm_phone(current_buyer.phone) == phone_key
        name_match = norm_name(current_buyer.name) == name_key
        if email_match or phone_match or name_match:
          buyer.tickets_cur = current_buyer.tickets_cur
          buyer.amount_cur = current_buyer.amount_cur
          if not buyer.email and current_buyer.email:
            buyer.email = current_buyer.email
          if not buyer.phone and current_buyer.phone:
            buyer.phone = current_buyer.phone
          break
      returning.append(buyer)
    else:
      prospects.append(buyer)

  new_buyers = [
    buyer for buyer in map_cur.values()
    if not (
      (norm_email(buyer.email) and norm_email(buyer.email) in emails_prev)
      or (norm_phone(buyer.phone) and norm_phone(buyer.phone) in phones_prev)
      or (norm_name(buyer.name) and norm_name(buyer.name) in names_prev)
    )
  ]

  returning.sort(key=lambda buyer: buyer.amount_cur, reverse=True)
  prospects.sort(key=lambda buyer: buyer.amount_prev, reverse=True)
  new_buyers.sort(key=lambda buyer: buyer.amount_cur, reverse=True)

  all_labels: list[str] = []
  cursor = date(2001, 1, 1)
  while cursor.year == 2001:
    all_labels.append(cursor.strftime("%m-%d"))
    cursor += timedelta(days=1)
  today_mmdd = TODAY.strftime("%m-%d")

  def build_cumul(daily: dict[str, float], year: int, labels: list[str], cap: str | None) -> list[float | None]:
    running = 0.0
    out: list[float | None] = []
    for mmdd in labels:
      if cap and mmdd > cap:
        out.append(None)
        continue
      running += daily.get(f"{year}-{mmdd}", 0.0)
      out.append(round(running, 2))
    return out

  event_end_mmdd = "04-11"
  prev_days_with_sales = [day[5:] for day, value in daily_prev.items() if value > 0]
  cur_days_with_sales = [day[5:] for day, value in daily_cur.items() if value > 0]
  window_start = min(
    min(prev_days_with_sales) if prev_days_with_sales else today_mmdd,
    min(cur_days_with_sales) if cur_days_with_sales else today_mmdd,
  )
  master_labels = [label for label in all_labels if window_start <= label <= event_end_mmdd]
  chart_prev = build_cumul(daily_prev, PREV_YEAR, master_labels, None)
  chart_cur = build_cumul(daily_cur, CURRENT_YEAR, master_labels, today_mmdd)

  def _email_link(email: str) -> str:
    escaped = esc(email)
    return f'<a href="mailto:{escaped}">{escaped}</a>' if escaped else ""

  def ret_row(buyer: Buyer) -> str:
    phone = esc(buyer.phone)
    return (
      f'<tr data-search="{esc(buyer.name)} {esc(buyer.email)} {phone}">'
      f'<td>{esc(buyer.name)}</td><td>{_email_link(buyer.email)}</td><td>{phone}</td>'
      f'<td class="r">{buyer.tickets_prev or "—"}</td>'
      f'<td class="r amt">${buyer.amount_prev:,.0f}</td>'
      f'<td class="r">{buyer.tickets_cur or "—"}</td>'
      f'<td class="r amt hi">${buyer.amount_cur:,.0f}</td>'
      f'</tr>'
    )

  def pro_row(buyer: Buyer) -> str:
    phone = esc(buyer.phone)
    return (
      f'<tr data-search="{esc(buyer.name)} {esc(buyer.email)} {phone}">'
      f'<td>{esc(buyer.name)}</td><td>{_email_link(buyer.email)}</td><td>{phone}</td>'
      f'<td class="r">{buyer.tickets_prev or "—"}</td>'
      f'<td class="r amt">${buyer.amount_prev:,.0f}</td>'
      f'</tr>'
    )

  def new_row(buyer: Buyer) -> str:
    phone = esc(buyer.phone)
    return (
      f'<tr data-search="{esc(buyer.name)} {esc(buyer.email)} {phone}">'
      f'<td>{esc(buyer.name)}</td><td>{_email_link(buyer.email)}</td><td>{phone}</td>'
      f'<td class="r">{buyer.tickets_cur or "—"}</td>'
      f'<td class="r amt hi">${buyer.amount_cur:,.0f}</td>'
      f'</tr>'
    )

  empty_ret = '<tr><td colspan="7" class="empty">No returning buyers yet.</td></tr>'
  empty_pro = f'<tr><td colspan="5" class="empty">All {PREV_YEAR} buyers are already in {CURRENT_YEAR}!</td></tr>'
  empty_new = f'<tr><td colspan="5" class="empty">No first-time {CURRENT_YEAR} buyers yet.</td></tr>'

  ret_rate = round(len(returning) / len(map_prev) * 100, 1) if map_prev else 0.0
  pct_diff_raw = round((total_raised - todate_prev_amt) / todate_prev_amt * 100, 1) if todate_prev_amt else 0.0
  goal = config["goal"]
  goal_pct = min(100, round(total_raised / goal * 100, 1)) if goal > 0 else 0.0
  goal_remaining = max(0.0, goal - total_raised) if goal > 0 else 0.0

  return {
    "slug": config["slug"],
    "raffle_key": raffle_key,
    "raffle_name": config["display_name"],
    "raffle_short_name": config["short_name"],
    "has_history": include_history,
    "total_raised": total_raised,
    "total_tickets": total_tickets,
    "buyer_count": len(buyers_cur),
    "daily_labels": sorted_days,
    "daily_values": [daily_totals[day] for day in sorted_days],
    "tier_labels": [tier[0] for tier in tiers],
    "tier_values": [tier[1] for tier in tiers],
    "top5": top5,
    "goal": goal,
    "pct": goal_pct,
    "goal_remaining": goal_remaining,
    "year": CURRENT_YEAR,
    "prev_year": PREV_YEAR,
    "today": TODAY.strftime("%B %d, %Y"),
    "updated_at": datetime.now(_TZ).strftime("%B %d, %Y %I:%M %p %Z"),
    "todate_prev_amt": todate_prev_amt,
    "todate_prev_tix": todate_prev_tix,
    "todate_prev_buyers": len(todate_prev_keys),
    "pct_diff": pct_diff_raw,
    "same_day_prev": SAME_DAY_PREV.strftime("%b %d"),
    "chart_master_labels": json.dumps(master_labels),
    "chart_prev": json.dumps(chart_prev),
    "chart_cur": json.dumps(chart_cur),
    "today_mmdd": json.dumps(today_mmdd),
    "n_returning": len(returning),
    "n_prospects": len(prospects),
    "n_new": len(new_buyers),
    "ret_rate": ret_rate,
    "rows_returning": "".join(ret_row(buyer) for buyer in returning) or empty_ret,
    "rows_prospects": "".join(pro_row(buyer) for buyer in prospects) or empty_pro,
    "rows_new": "".join(new_row(buyer) for buyer in new_buyers) or empty_new,
    "_returning": returning,
    "_prospects": prospects,
    "_new_buyers": new_buyers,
    "_top5": top5,
    "_daily_labels": sorted_days,
    "_daily_values": [daily_totals[day] for day in sorted_days],
  }

# ── Excel export ─────────────────────────────────────────────────────────────

def write_excel(d: dict, out_dir: str) -> str:
    """Write a multi-sheet Excel workbook and return its path."""
    wb = openpyxl.Workbook()

    # ── Shared styles ──────────────────────────────────────────────────────────
    DARK  = "1A0F35"
    CYAN  = "00B4CC"
    GOLD  = "C8A800"
    PURP  = "7C3AED"
    GREEN = "16A34A"
    WHITE = "F0E6FF"
    GREY  = "A78BCA"
    LIGHT = "2D1B5E"

    def hdr_font(color=WHITE):       return Font(bold=True, color=color, size=11)
    def cell_font(color=WHITE):      return Font(color=color, size=10)
    def hdr_fill(hex_color=DARK):    return PatternFill("solid", fgColor=hex_color)
    def thin_border():
        s = Side(style="thin", color="3D2870")
        return Border(left=s, right=s, top=s, bottom=s)
    def money(ws, cell, val):
        ws[cell] = val
        ws[cell].number_format = '$#,##0.00'
        ws[cell].font = cell_font(CYAN)
        ws[cell].alignment = Alignment(horizontal="right")
    def pct_cell(ws, cell, val):
        ws[cell] = val / 100
        ws[cell].number_format = '0.0%'
        ws[cell].font = cell_font(GOLD)
        ws[cell].alignment = Alignment(horizontal="right")

    def write_header_row(ws, headers: list[tuple[str,int,str]]):
        """headers = list of (label, col_width, hex_color)"""
        for col_idx, (label, width, color) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = hdr_font(WHITE)
            cell.fill = PatternFill("solid", fgColor=LIGHT)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    def style_data_row(ws, row: int, n_cols: int, alt: bool = False):
        bg = "1F1042" if alt else DARK
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            if not cell.font or cell.font.color.rgb in ("FF000000", "00000000"):
                cell.font = cell_font(WHITE)

    # ── Sheet 1: Summary ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_properties.tabColor = PURP
    ws1.sheet_view.showGridLines = False

    summary_rows = [
        ("Metric",                                  "Value"),
        (f"{d['year']} Total Raised",               f"${d['total_raised']:,.0f}"),
        (f"{d['year']} Tickets Sold",               d['total_tickets']),
        (f"{d['year']} Total Buyers",               d['buyer_count']),
        (f"Goal (${d['goal']:,})",                  f"{d['pct']}%"),
        ("", ""),
        (f"{d['prev_year']} Revenue (through {d['same_day_prev']})", f"${d['todate_prev_amt']:,.0f}"),
        (f"{d['prev_year']} Tickets (through {d['same_day_prev']})", d['todate_prev_tix']),
        (f"{d['prev_year']} Buyers (through {d['same_day_prev']})",  d['todate_prev_buyers']),
        (f"YoY Revenue Change (through {d['today']})", f"{'+' if d['pct_diff'] >= 0 else ''}{d['pct_diff']}%"),
        ("", ""),
        (f"Returning Buyers ({d['prev_year']} → {d['year']})", d['n_returning']),
        ("Retention Rate",                           f"{d['ret_rate']}%"),
        (f"Prospects (2025 buyers not yet in {d['year']})", d['n_prospects']),
        (f"New Buyers (first-timers in {d['year']})",d['n_new']),
        ("", ""),
        ("Report Generated",                         d['updated_at']),
    ]
    ws1.column_dimensions["A"].width = 46
    ws1.column_dimensions["B"].width = 22
    for r_idx, (label, value) in enumerate(summary_rows, 1):
        ca = ws1.cell(row=r_idx, column=1, value=label)
        cb = ws1.cell(row=r_idx, column=2, value=value)
        bg = LIGHT if r_idx == 1 else ("1F1042" if r_idx % 2 == 0 else DARK)
        for c in (ca, cb):
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center")
        if r_idx == 1:
            ca.font = hdr_font(); cb.font = hdr_font(CYAN)
        else:
            ca.font = cell_font(GREY); cb.font = cell_font(GOLD if str(value).endswith("%") else CYAN)

    # ── Sheet 2: Returning Buyers ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Returning Buyers")
    ws2.sheet_properties.tabColor = GREEN
    ws2.sheet_view.showGridLines = False
    headers2 = [
        ("Name", 30, WHITE),
        ("Email", 32, WHITE),
        ("Phone", 16, WHITE),
        (f"{d['prev_year']} Tickets", 14, WHITE),
        (f"{d['prev_year']} Paid", 14, WHITE),
        (f"{d['year']} Tickets", 14, WHITE),
        (f"{d['year']} Paid", 14, WHITE),
    ]
    write_header_row(ws2, headers2)
    for r_i, b in enumerate(d["_returning"], 2):
        ws2.cell(r_i, 1, b.name).font       = cell_font(WHITE)
        ws2.cell(r_i, 2, b.email).font      = cell_font(CYAN)
        ws2.cell(r_i, 3, b.phone).font      = cell_font(WHITE)
        ws2.cell(r_i, 4, b.tickets_prev or 0).font = cell_font(GREY)
        money(ws2, f"E{r_i}", b.amount_prev)
        ws2.cell(r_i, 6, b.tickets_cur or 0).font  = cell_font(WHITE)
        money(ws2, f"G{r_i}", b.amount_cur)
        ws2[f"G{r_i}"].font = cell_font(CYAN)
        style_data_row(ws2, r_i, 7, alt=(r_i % 2 == 0))

    # ── Sheet 3: Prospects ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Prospects")
    ws3.sheet_properties.tabColor = GOLD
    ws3.sheet_view.showGridLines = False
    headers3 = [
        ("Name", 30, WHITE),
        ("Email", 32, WHITE),
        ("Phone", 16, WHITE),
        (f"{d['prev_year']} Tickets", 14, WHITE),
        (f"{d['prev_year']} Paid", 14, WHITE),
    ]
    write_header_row(ws3, headers3)
    for r_i, b in enumerate(d["_prospects"], 2):
        ws3.cell(r_i, 1, b.name).font   = cell_font(WHITE)
        ws3.cell(r_i, 2, b.email).font  = cell_font(CYAN)
        ws3.cell(r_i, 3, b.phone).font  = cell_font(WHITE)
        ws3.cell(r_i, 4, b.tickets_prev or 0).font = cell_font(GREY)
        money(ws3, f"E{r_i}", b.amount_prev)
        style_data_row(ws3, r_i, 5, alt=(r_i % 2 == 0))

    # ── Sheet 4: New Buyers ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("New Buyers")
    ws4.sheet_properties.tabColor = "CC00AA"
    ws4.sheet_view.showGridLines = False
    headers4 = [
        ("Name", 30, WHITE),
        ("Email", 32, WHITE),
        ("Phone", 16, WHITE),
        (f"{d['year']} Tickets", 14, WHITE),
        (f"{d['year']} Paid", 14, WHITE),
    ]
    write_header_row(ws4, headers4)
    for r_i, b in enumerate(d["_new_buyers"], 2):
        ws4.cell(r_i, 1, b.name).font   = cell_font(WHITE)
        ws4.cell(r_i, 2, b.email).font  = cell_font(CYAN)
        ws4.cell(r_i, 3, b.phone).font  = cell_font(WHITE)
        ws4.cell(r_i, 4, b.tickets_cur or 0).font = cell_font(WHITE)
        money(ws4, f"E{r_i}", b.amount_cur)
        ws4[f"E{r_i}"].font = cell_font(CYAN)
        style_data_row(ws4, r_i, 5, alt=(r_i % 2 == 0))

    # ── Sheet 5: Daily Sales ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Daily Sales")
    ws5.sheet_properties.tabColor = CYAN
    ws5.sheet_view.showGridLines = False
    headers5 = [("Date", 14, WHITE), ("Daily Revenue", 18, WHITE), ("Cumulative Revenue", 20, WHITE)]
    write_header_row(ws5, headers5)
    running = 0.0
    for r_i, (day, val) in enumerate(zip(d["_daily_labels"], d["_daily_values"]), 2):
        running += val
        ws5.cell(r_i, 1, day).font = cell_font(GREY)
        money(ws5, f"B{r_i}", val)
        money(ws5, f"C{r_i}", running)
        ws5[f"C{r_i}"].font = cell_font(CYAN)
        style_data_row(ws5, r_i, 3, alt=(r_i % 2 == 0))

    out_path = os.path.join(out_dir, f"raffle_report_{d['year']}.xlsx")
    wb.save(out_path)
    return out_path


# ── HTML generation ───────────────────────────────────────────────────────────

def render_html(report_data: dict[str, dict]) -> str:
    raffles = [report_data[key] for key in RAFFLE_CONFIGS if key in report_data]
    if not raffles:
        raise ValueError("No raffle data available to render.")

    chart_payload = {
        raffle["slug"]: {
            "daily_labels": raffle["daily_labels"],
            "daily_values": raffle["daily_values"],
            "tier_labels": raffle["tier_labels"],
            "tier_values": raffle["tier_values"],
            "chart_master_labels": json.loads(raffle["chart_master_labels"]),
            "chart_prev": json.loads(raffle["chart_prev"]),
            "chart_cur": json.loads(raffle["chart_cur"]),
            "today_mmdd": json.loads(raffle["today_mmdd"]),
            "has_history": raffle["has_history"],
            "year": raffle["year"],
            "prev_year": raffle["prev_year"],
        }
        for raffle in raffles
    }

    def render_goal_block(raffle: dict) -> str:
        if raffle["goal"] <= 0:
            return (
                '<div class="goal-card goal-card--empty">'
                '<div class="goal-title">No goal configured</div>'
                '<div class="goal-empty-copy">Current totals are live, but this raffle does not have a progress target configured yet.</div>'
                '</div>'
            )

        return f"""
  <div class="goal-card">
    <div class="goal-header">
      <span class="goal-title">Goal Progress - ${raffle['goal']:,}</span>
      <span class="goal-pct">{raffle['pct']}%</span>
    </div>
    <div class="bar-track"><div class="bar-fill" style="width:{raffle['pct']}%"></div></div>
    <div class="goal-sub">
      <span>${raffle['total_raised']:,.0f} raised</span>
      <span>${raffle['goal_remaining']:,.0f} to go</span>
    </div>
  </div>"""

    def render_top5_rows(raffle: dict) -> str:
        rows = "".join(
            f'<tr><td class="td-name">{esc(name)}</td><td class="td-amount">${amt:,.0f}</td></tr>'
            for name, amt in raffle["top5"]
        )
        if rows:
            return rows
        return '<tr><td class="td-name" colspan="2" style="color:var(--muted);text-align:center">No data yet</td></tr>'

    def render_retention_panel(raffle: dict) -> str:
        if not raffle["has_history"]:
            return f"""
  <div id="tab-{raffle['slug']}-retention" class="tab-panel">
    <div class="section section-note">
      <div class="sec-hdr">
        <span class="sec-title">Current-Year Only</span>
      </div>
      <p class="sec-desc">No prior-year data exists yet for {esc(raffle['raffle_name'])}, so this page only shows live totals, ticket mix, and top buyers for {raffle['year']}.</p>
    </div>
  </div>"""

        arrow = "▲" if raffle["pct_diff"] >= 0 else "▼"
        arrow_cls = "up" if raffle["pct_diff"] >= 0 else "dn"
        badge_cls = "badge-up" if raffle["pct_diff"] >= 0 else "badge-dn"
        sign = "+" if raffle["pct_diff"] >= 0 else ""

        return f"""
  <div id="tab-{raffle['slug']}-retention" class="tab-panel">
    <div class="ret-grid">
      <div class="ret-card">
        <div class="ret-label">{raffle['year']} Buyers</div>
        <div class="ret-val cyan">{raffle['buyer_count']}</div>
      </div>
      <div class="ret-card">
        <div class="ret-label">{raffle['prev_year']} Buyers</div>
        <div class="ret-val purp">{raffle['n_returning'] + raffle['n_prospects']}</div>
      </div>
      <div class="ret-card">
        <div class="ret-label">Returning</div>
        <div class="ret-val grn">{raffle['n_returning']}</div>
        <div class="ret-sub">Retention: {raffle['ret_rate']}%</div>
      </div>
      <div class="ret-card">
        <div class="ret-label">Prospects</div>
        <div class="ret-val gold">{raffle['n_prospects']}</div>
        <div class="ret-sub">{raffle['prev_year']} buyers not yet in {raffle['year']}</div>
      </div>
      <div class="ret-card">
        <div class="ret-label">New Buyers</div>
        <div class="ret-val pink">{raffle['n_new']}</div>
        <div class="ret-sub">First-timers in {raffle['year']}</div>
      </div>
    </div>

    <div class="yoy-card">
      <div class="yoy-title">Year-over-Year - through {raffle['today']} vs same date in {raffle['prev_year']}</div>
      <div class="yoy-grid">
        <div class="yoy-col">
          <div class="yoy-yr">{raffle['prev_year']} (through {raffle['same_day_prev']})</div>
          <div class="yoy-val purp">${raffle['todate_prev_amt']:,.0f}</div>
          <div class="yoy-sub">{raffle['todate_prev_tix']:,} tickets · {raffle['todate_prev_buyers']} buyers</div>
        </div>
        <div class="vs-col">
          <div class="vs-arrow {arrow_cls}">{arrow}</div>
          <div class="yoy-badge {badge_cls}">{sign}{raffle['pct_diff']}%</div>
        </div>
        <div class="yoy-col">
          <div class="yoy-yr">{raffle['year']} (to date)</div>
          <div class="yoy-val cyan">${raffle['total_raised']:,.0f}</div>
          <div class="yoy-sub">{raffle['total_tickets']:,} tickets · {raffle['buyer_count']} buyers</div>
        </div>
      </div>
    </div>

    <div class="chart-card chart-card-line">
      <div class="chart-title">Cumulative Sales - {raffle['prev_year']} (full year) vs {raffle['year']} (to date)</div>
      <canvas id="{raffle['slug']}-lineChart" style="max-height:360px"></canvas>
    </div>

    <div class="section">
      <div class="sec-hdr">
        <span class="sec-title">Returning Buyers</span>
        <span class="badge-count">{raffle['n_returning']}</span>
      </div>
      <p class="sec-desc">Bought tickets in both {raffle['prev_year']} and {raffle['year']}.</p>
      <div class="search-wrap"><input type="text" placeholder="Search name, email, phone..." oninput="filterTable('tbl-{raffle['slug']}-ret', this.value)" /></div>
      <div class="tbl-wrap">
        <table id="tbl-{raffle['slug']}-ret" class="ret-tbl">
          <thead><tr>
            <th>Name</th><th>Email</th><th>Phone</th>
            <th class="r">{raffle['prev_year']} Tix</th><th class="r">{raffle['prev_year']} Paid</th>
            <th class="r">{raffle['year']} Tix</th><th class="r">{raffle['year']} Paid</th>
          </tr></thead>
          <tbody>{raffle['rows_returning']}</tbody>
        </table>
      </div>
    </div>

    <div class="section">
      <div class="sec-hdr">
        <span class="sec-title">Prospects</span>
        <span class="badge-count">{raffle['n_prospects']}</span>
      </div>
      <p class="sec-desc">Bought in {raffle['prev_year']} but not yet seen in {raffle['year']}.</p>
      <div class="search-wrap"><input type="text" placeholder="Search name, email, phone..." oninput="filterTable('tbl-{raffle['slug']}-pro', this.value)" /></div>
      <div class="tbl-wrap">
        <table id="tbl-{raffle['slug']}-pro" class="ret-tbl">
          <thead><tr>
            <th>Name</th><th>Email</th><th>Phone</th>
            <th class="r">{raffle['prev_year']} Tix</th><th class="r">{raffle['prev_year']} Paid</th>
          </tr></thead>
          <tbody>{raffle['rows_prospects']}</tbody>
        </table>
      </div>
    </div>

    <div class="section">
      <div class="sec-hdr">
        <span class="sec-title">New {raffle['year']} Buyers</span>
        <span class="badge-count">{raffle['n_new']}</span>
      </div>
      <p class="sec-desc">Purchased in {raffle['year']} with no record found in {raffle['prev_year']}.</p>
      <div class="search-wrap"><input type="text" placeholder="Search name, email, phone..." oninput="filterTable('tbl-{raffle['slug']}-new', this.value)" /></div>
      <div class="tbl-wrap">
        <table id="tbl-{raffle['slug']}-new" class="ret-tbl">
          <thead><tr>
            <th>Name</th><th>Email</th><th>Phone</th>
            <th class="r">{raffle['year']} Tix</th><th class="r">{raffle['year']} Paid</th>
          </tr></thead>
          <tbody>{raffle['rows_new']}</tbody>
        </table>
      </div>
    </div>
  </div>"""

    def render_raffle_page(raffle: dict, active: bool) -> str:
        history_button = ""
        if raffle["has_history"]:
            history_button = f'<button class="tab-btn" onclick="switchTab(\'{raffle["slug"]}\', \'retention\', this)">Retention &amp; YoY</button>'

        dashboard_note = ""
        if not raffle["has_history"]:
            dashboard_note = (
                '<div class="section section-note">'
                '<div class="sec-hdr"><span class="sec-title">Fresh Raffle</span></div>'
                f'<p class="sec-desc">{esc(raffle["raffle_name"])} is running without prior-year comparison data, so this page focuses on live totals and sales mix only.</p>'
                '</div>'
            )

        return f"""
  <section class="raffle-page{' active' if active else ''}" data-raffle-page="{raffle['slug']}">
    <div class="raffle-hero">
      <div>
        <div class="eyebrow">Live Raffle View</div>
        <h2>{esc(raffle['raffle_name'])}</h2>
        <p>{raffle['year']} totals and sales details{'' if raffle['has_history'] else ' with current-year reporting only'}.</p>
      </div>
      <div class="hero-stat">
        <span class="hero-stat-label">Tickets Sold</span>
        <span class="hero-stat-value">{raffle['total_tickets']}</span>
      </div>
    </div>

    <div class="tabs tabs-inner">
      <button class="tab-btn active" onclick="switchTab('{raffle['slug']}', 'dashboard', this)">Dashboard</button>
      {history_button}
    </div>

    <div id="tab-{raffle['slug']}-dashboard" class="tab-panel active">
      <div class="grid">
        <div class="card">
          <div class="card-label">Total Raised</div>
          <div class="card-value gold">${raffle['total_raised']:,.0f}</div>
        </div>
        <div class="card">
          <div class="card-label">Tickets Sold</div>
          <div class="card-value cyan">{raffle['total_tickets']}</div>
        </div>
        <div class="card">
          <div class="card-label">Total Buyers</div>
          <div class="card-value purp">{raffle['buyer_count']}</div>
        </div>
      </div>

      {render_goal_block(raffle)}

      <div class="charts-row">
        <div class="chart-card">
          <div class="chart-title">Daily Sales ($)</div>
          <canvas id="{raffle['slug']}-barChart"></canvas>
        </div>
        <div class="chart-card">
          <div class="chart-title">Ticket Tier Mix</div>
          <canvas id="{raffle['slug']}-donutChart"></canvas>
        </div>
      </div>

      <div class="leaderboard">
        <div class="lb-title">Top Supporters</div>
        <table>
          <tbody>{render_top5_rows(raffle)}</tbody>
        </table>
      </div>

      {dashboard_note}
    </div>

    {render_retention_panel(raffle)}
  </section>"""

    def render_raffle_switch(raffle: dict, active: bool) -> str:
      active_class = " active" if active else ""
      return (
        f'<button class="raffle-switch{active_class}" data-raffle-btn="{raffle["slug"]}" '
        f'onclick="switchRaffle(\'{raffle["slug"]}\', this)">'
        f'{esc(raffle["raffle_short_name"])}'
        f'<span>{esc(raffle["raffle_name"])} totals</span>'
        '</button>'
      )

    raffle_switcher = "".join(
      render_raffle_switch(raffle, idx == 0)
      for idx, raffle in enumerate(raffles)
    )
    raffle_pages = "".join(render_raffle_page(raffle, idx == 0) for idx, raffle in enumerate(raffles))

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Sir Pork-a-Lot · Multi-Raffle Dashboard {raffles[0]['year']}</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
  <style>
    :root {{
      --bg:      #0f0820;
      --card:    #1a0f35;
      --cyan:    #00e5ff;
      --pink:    #ff00cc;
      --gold:    #ffd700;
      --purple:  #c084fc;
      --grn:     #4ade80;
      --red:     #f87171;
      --text:    #f0e6ff;
      --muted:   #a78bca;
      --border:  rgba(192,132,252,.22);
    }}
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      background: radial-gradient(ellipse at 50% 0%, rgba(160,80,255,.22) 0%, transparent 55%),
                  linear-gradient(180deg, #0f0820 0%, #1a0d35 100%);
      min-height: 100vh;
      color: var(--text);
      font-family: Arial, Helvetica, sans-serif;
      padding: 2rem 1rem 4rem;
    }}
    /* ── Header ── */
    .header {{
      text-align: center;
      margin-bottom: 2.5rem;
    }}
    .header img {{
      width: 140px;
      mix-blend-mode: screen;
      filter: drop-shadow(0 0 28px rgba(255,0,204,.65));
      margin-bottom: .75rem;
    }}
    .header h1 {{
      font-size: 1.5rem;
      font-weight: 800;
      letter-spacing: .12em;
      text-transform: uppercase;
      color: var(--purple);
    }}
    .header .sub {{
      font-size: .9rem;
      color: var(--muted);
      margin-top: .25rem;
    }}
    .divider {{
      height: 1px;
      background: linear-gradient(90deg, transparent, var(--pink), var(--cyan), var(--pink), transparent);
      box-shadow: 0 0 8px rgba(255,0,204,.5);
      max-width: 360px;
      margin: .75rem auto 0;
    }}
    .raffle-switcher {{
      max-width: 960px;
      margin: 0 auto 1.5rem;
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: .9rem;
    }}
    .raffle-switch {{
      display: flex;
      flex-direction: column;
      align-items: flex-start;
      gap: .15rem;
      background: rgba(26,15,53,.78);
      border: 1px solid rgba(192,132,252,.2);
      border-radius: 1rem;
      color: var(--text);
      padding: 1rem 1.1rem;
      cursor: pointer;
      transition: border-color .18s, transform .18s, box-shadow .18s;
    }}
    .raffle-switch span {{
      color: var(--muted);
      font-size: .74rem;
      letter-spacing: .04em;
      text-transform: uppercase;
    }}
    .raffle-switch:hover {{
      transform: translateY(-1px);
      border-color: var(--cyan);
    }}
    .raffle-switch.active {{
      border-color: var(--cyan);
      box-shadow: 0 0 0 1px rgba(0,229,255,.12), 0 0 18px rgba(0,229,255,.12);
      background: rgba(0,229,255,.08);
    }}
    .raffle-page {{ display: none; }}
    .raffle-page.active {{ display: block; }}
    .raffle-hero {{
      max-width: 960px;
      margin: 0 auto 1.25rem;
      display: flex;
      justify-content: space-between;
      align-items: flex-end;
      gap: 1rem;
      padding: 1.4rem 1.55rem;
      background: linear-gradient(135deg, rgba(255,0,204,.09), rgba(0,229,255,.08));
      border: 1px solid rgba(192,132,252,.22);
      border-radius: 1rem;
    }}
    .raffle-hero h2 {{
      font-size: 1.45rem;
      letter-spacing: .04em;
      margin-bottom: .25rem;
    }}
    .raffle-hero p {{ color: var(--muted); font-size: .88rem; }}
    .eyebrow {{
      color: var(--cyan);
      text-transform: uppercase;
      letter-spacing: .1em;
      font-size: .68rem;
      margin-bottom: .35rem;
    }}
    .hero-stat {{ text-align: right; min-width: 140px; }}
    .hero-stat-label {{
      display: block;
      color: var(--muted);
      font-size: .68rem;
      text-transform: uppercase;
      letter-spacing: .08em;
      margin-bottom: .18rem;
    }}
    .hero-stat-value {{ font-size: 2rem; font-weight: 800; color: var(--gold); }}
    /* ── Grid ── */
    .grid {{
      max-width: 960px;
      margin: 0 auto;
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
      gap: 1.25rem;
      margin-bottom: 1.5rem;
    }}
    .card {{
      background: rgba(26,15,53,.75);
      border: 1px solid rgba(192,132,252,.25);
      border-radius: 1rem;
      padding: 1.4rem 1.5rem;
    }}
    .card-label {{
      font-size: .75rem;
      text-transform: uppercase;
      letter-spacing: .08em;
      color: var(--muted);
      margin-bottom: .4rem;
    }}
    .card-value {{
      font-size: 2rem;
      font-weight: 800;
      line-height: 1.1;
    }}
    .cyan  {{ color: var(--cyan);   text-shadow: 0 0 12px rgba(0,229,255,.6); }}
    .gold  {{ color: var(--gold);   text-shadow: 0 0 12px rgba(255,215,0,.6); }}
    .pink  {{ color: var(--pink);   text-shadow: 0 0 12px rgba(255,0,204,.6); }}
    .purp  {{ color: var(--purple); text-shadow: 0 0 12px rgba(192,132,252,.6); }}
    /* ── Goal bar ── */
    .goal-card {{
      max-width: 960px;
      margin: 0 auto 1.5rem;
      background: rgba(26,15,53,.75);
      border: 1px solid rgba(255,215,0,.3);
      border-radius: 1rem;
      padding: 1.4rem 1.75rem;
    }}
    .goal-header {{
      display: flex;
      justify-content: space-between;
      align-items: baseline;
      margin-bottom: .75rem;
    }}
    .goal-title {{
      font-size: 1rem;
      font-weight: 700;
      color: var(--gold);
      text-shadow: 0 0 10px rgba(255,215,0,.6);
    }}
    .goal-pct {{
      font-size: 1.4rem;
      font-weight: 800;
      color: var(--gold);
    }}
    .bar-track {{
      background: rgba(255,255,255,.07);
      border-radius: 9999px;
      height: 18px;
      overflow: hidden;
    }}
    .bar-fill {{
      height: 100%;
      border-radius: 9999px;
      background: linear-gradient(90deg, #cc00aa, #ff00cc, #ffd700);
      box-shadow: 0 0 16px rgba(255,0,204,.5);
      transition: width .6s ease;
      width: 0;
    }}
    .goal-sub {{
      display: flex;
      justify-content: space-between;
      font-size: .8rem;
      color: var(--muted);
      margin-top: .5rem;
    }}
    .goal-card--empty {{ border-color: rgba(0,229,255,.22); }}
    .goal-empty-copy {{ color: var(--muted); font-size: .88rem; margin-top: .45rem; }}
    /* ── Charts row ── */
    .charts-row {{
      max-width: 960px;
      margin: 0 auto 1.5rem;
      display: grid;
      grid-template-columns: 2fr 1fr;
      gap: 1.25rem;
    }}
    @media (max-width: 640px) {{
      .charts-row {{ grid-template-columns: 1fr; }}
    }}
    .chart-card {{
      background: rgba(26,15,53,.75);
      border: 1px solid rgba(192,132,252,.25);
      border-radius: 1rem;
      padding: 1.25rem;
    }}
    .chart-title {{
      font-size: .8rem;
      text-transform: uppercase;
      letter-spacing: .08em;
      color: var(--muted);
      margin-bottom: 1rem;
    }}
    /* ── Top buyers ── */
    .leaderboard {{
      max-width: 960px;
      margin: 0 auto 1.5rem;
      background: rgba(26,15,53,.75);
      border: 1px solid rgba(192,132,252,.25);
      border-radius: 1rem;
      padding: 1.4rem 1.75rem;
    }}
    .lb-title {{
      font-size: .8rem;
      text-transform: uppercase;
      letter-spacing: .08em;
      color: var(--muted);
      margin-bottom: 1rem;
    }}
    table {{ width: 100%; border-collapse: collapse; }}
    tr:not(:last-child) td {{ border-bottom: 1px solid rgba(192,132,252,.12); }}
    td {{ padding: .6rem .25rem; font-size: .95rem; }}
    .td-name   {{ color: var(--text); }}
    .td-amount {{ text-align: right; font-weight: 700; color: var(--cyan);
                  text-shadow: 0 0 8px rgba(0,229,255,.4); }}
    /* ── Tabs ── */
    .tabs {{
      max-width: 960px;
      margin: 0 auto 1.25rem;
      display: flex;
      gap: .5rem;
      flex-wrap: wrap;
    }}
    .tab-btn {{
      background: rgba(26,15,53,.8);
      border: 1px solid rgba(192,132,252,.25);
      border-radius: .6rem;
      color: var(--muted);
      font-size: .8rem;
      font-weight: 700;
      letter-spacing: .06em;
      padding: .45rem 1rem;
      cursor: pointer;
      transition: border-color .15s, color .15s;
    }}
    .tab-btn:hover {{ border-color: var(--cyan); color: var(--cyan); }}
    .tab-btn.active {{
      border-color: var(--cyan);
      color: var(--cyan);
      background: rgba(0,229,255,.08);
      text-shadow: 0 0 8px rgba(0,229,255,.4);
    }}
    .tab-panel {{ display: none; }}
    .tab-panel.active {{ display: block; }}
    .tabs-inner {{ margin-bottom: 1.25rem; }}
    /* ── YoY card ── */
    .yoy-card {{
      max-width: 960px;
      margin: 0 auto 1.5rem;
      background: rgba(26,15,53,.8);
      border: 1px solid rgba(255,215,0,.28);
      border-radius: 1rem;
      padding: 1.4rem 1.75rem;
    }}
    .yoy-title {{
      font-size: .85rem; font-weight: 700;
      color: var(--gold); text-transform: uppercase;
      letter-spacing: .1em; margin-bottom: 1rem;
      text-shadow: 0 0 8px rgba(255,215,0,.5);
    }}
    .yoy-grid {{
      display: grid;
      grid-template-columns: 1fr auto 1fr;
      gap: .75rem;
      align-items: center;
    }}
    .yoy-col {{ text-align: center; }}
    .yoy-yr  {{ font-size: .7rem; text-transform: uppercase; letter-spacing: .1em; color: var(--muted); margin-bottom: .2rem; }}
    .yoy-val {{ font-size: 1.55rem; font-weight: 800; }}
    .yoy-sub {{ font-size: .72rem; color: var(--muted); margin-top: .15rem; }}
    .vs-col  {{ text-align: center; }}
    .vs-arrow {{ font-size: 1.5rem; font-weight: 900; }}
    .yoy-badge {{
      display: inline-block; font-size: .8rem; font-weight: 700;
      padding: .15rem .55rem; border-radius: 999px; margin-top: .3rem;
    }}
    .badge-up {{ background: rgba(74,222,128,.15); color: var(--grn); border: 1px solid rgba(74,222,128,.3); }}
    .badge-dn {{ background: rgba(248,113,113,.12); color: var(--red); border: 1px solid rgba(248,113,113,.3); }}
    .up {{ color: var(--grn); text-shadow: 0 0 8px rgba(74,222,128,.4); }}
    .dn {{ color: var(--red); text-shadow: 0 0 8px rgba(248,113,113,.4); }}
    /* ── Retention stat chips ── */
    .ret-grid {{
      max-width: 960px;
      margin: 0 auto 1.25rem;
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
      gap: .9rem;
    }}
    .ret-card {{
      background: rgba(26,15,53,.8);
      border: 1px solid var(--border);
      border-radius: 1rem;
      padding: 1.1rem 1.25rem;
    }}
    .ret-label {{ font-size: .7rem; text-transform: uppercase; letter-spacing: .08em; color: var(--muted); margin-bottom: .3rem; }}
    .ret-val   {{ font-size: 1.7rem; font-weight: 800; line-height: 1.1; }}
    .ret-sub   {{ font-size: .68rem; color: var(--muted); margin-top: .25rem; }}
    .grn {{ color: var(--grn); text-shadow: 0 0 10px rgba(74,222,128,.55); }}
    /* ── Section panels ── */
    .section {{
      max-width: 960px;
      margin: 0 auto 1.5rem;
      background: rgba(26,15,53,.8);
      border: 1px solid var(--border);
      border-radius: 1rem;
      padding: 1.4rem 1.75rem;
    }}
    .sec-hdr {{
      display: flex; align-items: baseline; gap: .6rem; margin-bottom: .75rem;
    }}
    .sec-title {{
      font-size: .78rem; text-transform: uppercase;
      letter-spacing: .1em; color: var(--muted);
    }}
    .badge-count {{
      font-size: .72rem; font-weight: 700;
      padding: .1rem .45rem; border-radius: 999px;
      background: rgba(192,132,252,.18);
      color: var(--purple); border: 1px solid rgba(192,132,252,.3);
    }}
    .sec-desc {{ font-size: .78rem; color: var(--muted); margin-bottom: .85rem; }}
    .search-wrap {{ margin-bottom: .85rem; }}
    .search-wrap input {{
      width: 100%; max-width: 360px;
      background: rgba(255,255,255,.05);
      border: 1px solid rgba(192,132,252,.3);
      border-radius: .5rem;
      color: var(--text); font-size: .85rem;
      padding: .38rem .7rem; outline: none;
    }}
    .search-wrap input::placeholder {{ color: var(--muted); }}
    .search-wrap input:focus {{ border-color: var(--cyan); }}
    /* ── Retention tables ── */
    .tbl-wrap {{ overflow-x: auto; }}
    .ret-tbl {{ width: 100%; border-collapse: collapse; font-size: .855rem; min-width: 480px; }}
    .ret-tbl thead tr {{ border-bottom: 1px solid rgba(192,132,252,.3); }}
    .ret-tbl th {{
      padding: .45rem .45rem; text-align: left;
      font-size: .68rem; text-transform: uppercase;
      letter-spacing: .08em; color: var(--muted); white-space: nowrap;
    }}
    .ret-tbl th.r, .ret-tbl td.r {{ text-align: right; }}
    .ret-tbl tbody tr {{ border-bottom: 1px solid rgba(192,132,252,.08); transition: background .15s; }}
    .ret-tbl tbody tr:hover {{ background: rgba(192,132,252,.08); }}
    .ret-tbl tbody tr.hidden {{ display: none; }}
    .ret-tbl td {{ padding: .5rem .45rem; }}
    .ret-tbl td a {{ color: var(--cyan); text-decoration: none; font-size: .78rem; }}
    .ret-tbl td a:hover {{ text-decoration: underline; }}
    .ret-tbl td.amt {{ font-weight: 700; color: var(--muted); }}
    .ret-tbl td.hi  {{ color: var(--cyan); text-shadow: 0 0 8px rgba(0,229,255,.4); font-weight: 700; }}
    .ret-tbl td.empty {{ text-align: center; color: var(--muted); padding: 1.5rem; font-style: italic; }}
    .section-note {{ border-style: dashed; }}
    .chart-card-line {{ max-width:960px; margin:0 auto 1.5rem; }}
    /* ── Footer ── */
    .footer {{
      text-align: center;
      margin-top: 2rem;
      font-size: .75rem;
      color: rgba(167,139,202,.4);
    }}
    @media (max-width: 640px) {{
      .raffle-hero {{ flex-direction: column; align-items: flex-start; }}
      .hero-stat {{ text-align: left; }}
      .goal-sub {{ flex-direction: column; gap: .25rem; }}
    }}
  </style>
</head>
<body>

  <div class="header">
    <img src="https://sirporkalot.vercel.app/Photo_2.jpeg" alt="Sir Pork-a-Lot" />
    <h1>Raffle Dashboard {raffles[0]['year']}</h1>
    <p class="sub">Multi-raffle totals for Hogs for the Cause and APEX CRNA Essential Bundle · Last updated {raffles[0]['updated_at']}</p>
    <div class="divider"></div>
  </div>

  <div class="raffle-switcher">
    {raffle_switcher}
  </div>

  {raffle_pages}

  <div class="footer">&copy; {raffles[0]['year']} Team Sir Pork a Lot - unified raffle reporting</div>

  <script>
    const CYAN   = "#00e5ff";
    const PINK   = "#ff00cc";
    const GOLD   = "#ffd700";
    const PURPLE = "#c084fc";
    const GREEN  = "#4ade80";
    const chartData = {json.dumps(chart_payload)};
    const chartRegistry = {{}};

    function ensureDashboardCharts(slug) {{
      if (!chartRegistry[slug]) chartRegistry[slug] = {{}};
      const cfg = chartData[slug];
      if (!cfg) return;

      if (!chartRegistry[slug].bar) {{
        const barCanvas = document.getElementById(slug + '-barChart');
        if (barCanvas) {{
          chartRegistry[slug].bar = new Chart(barCanvas, {{
            type: 'bar',
            data: {{
              labels: cfg.daily_labels,
              datasets: [{{
                label: 'Daily Sales ($)',
                data: cfg.daily_values,
                backgroundColor: 'rgba(0,229,255,.25)',
                borderColor: CYAN,
                borderWidth: 2,
                borderRadius: 6,
              }}]
            }},
            options: {{
              responsive: true,
              plugins: {{ legend: {{ display: false }} }},
              scales: {{
                x: {{ ticks: {{ color: '#a78bca', maxRotation: 45 }}, grid: {{ color: 'rgba(192,132,252,.1)' }} }},
                y: {{ ticks: {{ color: '#a78bca', callback: v => '$' + v }}, grid: {{ color: 'rgba(192,132,252,.1)' }} }}
              }}
            }}
          }});
        }}
      }}

      if (!chartRegistry[slug].donut) {{
        const donutCanvas = document.getElementById(slug + '-donutChart');
        if (donutCanvas) {{
          chartRegistry[slug].donut = new Chart(donutCanvas, {{
            type: 'doughnut',
            data: {{
              labels: cfg.tier_labels,
              datasets: [{{
                data: cfg.tier_values,
                backgroundColor: [CYAN, PINK, GOLD, PURPLE],
                borderColor: '#0f0820',
                borderWidth: 3,
                hoverOffset: 8,
              }}]
            }},
            options: {{
              responsive: true,
              plugins: {{
                legend: {{ position: 'bottom', labels: {{ color: '#a78bca', padding: 14, font: {{ size: 11 }} }} }},
              }}
            }}
          }});
        }}
      }}

      Object.values(chartRegistry[slug]).forEach(chart => chart && chart.resize());
    }}

    function ensureRetentionCharts(slug) {{
      if (!chartRegistry[slug]) chartRegistry[slug] = {{}};
      const cfg = chartData[slug];
      if (!cfg || !cfg.has_history || chartRegistry[slug].line) return;

      const canvas = document.getElementById(slug + '-lineChart');
      if (!canvas) return;

      const tickLabels = cfg.chart_master_labels.map((label, index) => (index % 14 === 0 ? label : ''));
      chartRegistry[slug].line = new Chart(canvas, {{
        type: 'line',
        data: {{
          labels: cfg.chart_master_labels,
          datasets: [
            {{
              label: cfg.prev_year + ' Cumulative ($)',
              data: cfg.chart_prev,
              borderColor: PURPLE,
              backgroundColor: 'rgba(192,132,252,.07)',
              borderWidth: 2,
              pointRadius: 0,
              tension: 0.3,
              fill: false,
              spanGaps: false,
            }},
            {{
              label: cfg.year + ' Cumulative ($)',
              data: cfg.chart_cur,
              borderColor: CYAN,
              backgroundColor: 'rgba(0,229,255,.07)',
              borderWidth: 2.5,
              pointRadius: 0,
              tension: 0.3,
              fill: false,
              spanGaps: false,
            }},
            {{
              label: 'Today',
              data: cfg.chart_master_labels.map((label, index) => {{
                if (label !== cfg.today_mmdd) return null;
                return Math.max(cfg.chart_prev[index] || 0, cfg.chart_cur[index] || 0) * 1.05;
              }}),
              borderColor: 'rgba(255,215,0,.7)',
              borderWidth: 1.5,
              borderDash: [4, 4],
              pointRadius: 6,
              pointStyle: 'line',
              pointBorderColor: GOLD,
              showLine: false,
              spanGaps: false,
            }},
          ],
        }},
        options: {{
          responsive: true,
          interaction: {{ mode: 'index', intersect: false }},
          plugins: {{
            legend: {{ position: 'top', labels: {{ color: '#a78bca', font: {{ size: 11 }}, padding: 16 }} }},
            tooltip: {{ callbacks: {{ label: ctx => ctx.dataset.label + ': $' + (ctx.parsed.y || 0).toLocaleString() }} }},
          }},
          scales: {{
            x: {{ ticks: {{ color: '#a78bca', maxRotation: 45, callback: (value, index) => tickLabels[index] }}, grid: {{ color: 'rgba(192,132,252,.08)' }} }},
            y: {{ ticks: {{ color: '#a78bca', callback: value => '$' + value.toLocaleString() }}, grid: {{ color: 'rgba(192,132,252,.08)' }} }},
          }},
        }}
      }});
      chartRegistry[slug].line.resize();
    }}

    function switchRaffle(slug, btn) {{
      document.querySelectorAll('[data-raffle-page]').forEach(page => page.classList.remove('active'));
      document.querySelectorAll('[data-raffle-btn]').forEach(button => button.classList.remove('active'));
      document.querySelector('[data-raffle-page="' + slug + '"]').classList.add('active');
      btn.classList.add('active');
      ensureDashboardCharts(slug);
      const activeTab = document.querySelector('[data-raffle-page="' + slug + '"] .tab-panel.active');
      if (activeTab && activeTab.id.endsWith('-retention')) ensureRetentionCharts(slug);
    }}

    function switchTab(slug, name, btn) {{
      const page = document.querySelector('[data-raffle-page="' + slug + '"]');
      page.querySelectorAll('.tab-panel').forEach(panel => panel.classList.remove('active'));
      page.querySelectorAll('.tab-btn').forEach(button => button.classList.remove('active'));
      page.querySelector('#tab-' + slug + '-' + name).classList.add('active');
      btn.classList.add('active');
      if (name === 'dashboard') ensureDashboardCharts(slug);
      if (name === 'retention') ensureRetentionCharts(slug);
    }}

    function filterTable(tableId, query) {{
      const q = query.toLowerCase().trim();
      document.querySelectorAll('#' + tableId + ' tbody tr').forEach(tr => {{
        const s = (tr.dataset.search || tr.innerText).toLowerCase();
        tr.classList.toggle('hidden', q.length > 0 && !s.includes(q));
      }});
    }}

    document.addEventListener('DOMContentLoaded', () => {{
      const firstSlug = {json.dumps(raffles[0]['slug'])};
      ensureDashboardCharts(firstSlug);
    }});
  </script>
</body>
</html>"""

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Authenticating to Microsoft Graph...")
    token = get_token()

    print("Resolving SharePoint site ID...")
    site_id = get_site_id(token)
    print(f"  Site ID: {site_id}")

    print("Fetching all list items (all years)...")
    all_items = get_list_items(token, site_id)

    unique_raffles = sorted({str((item.get("fields") or {}).get("RaffleName") or "") for item in all_items})
    print(f"  Unique RaffleName values: {unique_raffles}")

    print(f"Loading {PREV_YEAR} CSV...")
    csv_rows = load_csv_prev()

    print("Processing raffle data...")
    hogs_data = process_raffle(all_items, "hogs_for_the_cause", csv_rows)
    crna_data = process_raffle(all_items, "crna_essential_bundle")
    report_data = {
      "hogs_for_the_cause": hogs_data,
      "crna_essential_bundle": crna_data,
    }
    print(f"  Hogs: ${hogs_data['total_raised']:,.0f} raised, {hogs_data['total_tickets']} tickets")
    print(f"  CRNA: ${crna_data['total_raised']:,.0f} raised, {crna_data['total_tickets']} tickets")

    print("Rendering HTML...")
    html = render_html(report_data)

    out_dir = os.path.join(os.path.dirname(__file__), "dist")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "index.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  Saved to {out_path}")

    print("Writing Excel report...")
    xlsx_path = write_excel(hogs_data, out_dir)
    print(f"  Saved to {xlsx_path}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
